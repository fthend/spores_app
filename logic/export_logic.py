import pandas as pd
from PySide6.QtWidgets import QFileDialog, QMessageBox

def process_stratigraphy(obj, fields):
    row = {}

    if not hasattr(obj, 'stratigraphic_periods') or not obj.stratigraphic_periods:
        for field in ["Период", "Эпоха", "Ярус"]:
            if field in fields:
                row[field] = ""
        return row

    include_period = "Период" in fields
    include_epoch = "Эпоха" in fields
    include_stage = "Ярус" in fields

    if include_period and include_epoch and include_stage:
        combined = []
        for sp in obj.stratigraphic_periods:
            parts = []
            if sp.period:
                parts.append(sp.period)
            if sp.epoch:
                parts.append(sp.epoch)
            full = " ".join(parts)
            if sp.stage:
                full += f", {sp.stage}"
            combined.append(full)
        row["Период"] = "; ".join(combined)
        row["Эпоха"] = ""
        row["Ярус"] = ""
    else:
        periods = []
        epochs = []
        stages = []

        for sp in obj.stratigraphic_periods:
            if include_period and sp.period:
                periods.append(sp.period)
            if include_epoch and sp.epoch:
                epochs.append(sp.epoch)
            if include_stage and sp.stage:
                stages.append(sp.stage)

        if include_period:
            row["Период"] = "; ".join(periods) if periods else ""
        if include_epoch:
            row["Эпоха"] = "; ".join(epochs) if epochs else ""
        if include_stage:
            row["Ярус"] = "; ".join(stages) if stages else ""

    return row


def process_geography(obj, fields):
    result = {}
    if not hasattr(obj, 'geographic_locations') or not obj.geographic_locations:
        return result

    pairs = []
    regions = []
    for loc in obj.geographic_locations:
        regions.append(loc.name)
        if loc.parent:
            pairs.append(f"{loc.parent.name}: {loc.name}")
        else:
            pairs.append(loc.name)

    if "Страна и регион" in fields:
        result["Страна и регион"] = ", ".join(pairs) if pairs else ""
    if "Регион" in fields:
        result["Регион"] = ", ".join(regions) if regions else ""

    return result

def save_to_file(df, export_format):
    file_types = {
        "CSV": "CSV Files (*.csv)",
        "Excel (XLSX)": "Excel Files (*.xlsx)",
        "JSON": "JSON Files (*.json)",
        "HTML": "HTML Files (*.html)"
    }

    file_dialog = QFileDialog()
    file_path, _ = file_dialog.getSaveFileName(
        None, "Сохранить файл", "", file_types[export_format]
    )

    if not file_path:
        return False

    try:
        if export_format == "CSV":
            # df.to_csv(file_path, index=False, encoding='utf-8-sig', sep=';')
            df = df.copy()
            for col in df.columns:
                if "Длина лучей щели" in col:
                    df[col] = df[col].apply(lambda x: f"\t{x}" if x and "/" in str(x) else x)
            df.to_csv(file_path, index=False, encoding='utf-8-sig', sep=';')
        elif export_format == "Excel (XLSX)":
            # df.to_excel(file_path, index=False, engine='openpyxl')
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active

            # Записываем заголовки
            for col_num, column_title in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)
                if "Длина лучей щели" in column_title:
                    # Устанавливаем текстовый формат для всего столбца
                    for cell in ws.iter_cols(min_col=col_num, max_col=col_num):
                        for x in cell:
                            x.number_format = '@'  # Текстовый формат

            # Записываем данные
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    if "Длина лучей щели" in df.columns[col_num - 1]:
                        # Для дробей добавляем апостроф в начале
                        if isinstance(value, str) and "/" in value:
                            value = "'" + value
                    ws.cell(row=row_idx, column=col_num, value=value)

            wb.save(file_path)
        elif export_format == "JSON":
            df.to_json(file_path, orient="records", indent=4, force_ascii=False)
        elif export_format == "HTML":
            df.to_html(file_path, index=False)

        QMessageBox.information(None, "Экспорт завершён", f"Файл успешно сохранён:\n{file_path}")
        return True
    except Exception as e:
        QMessageBox.critical(None, "Ошибка при экспорте", str(e))
        return False


# Преобразует объект Genus в словарь с учетом выбранных полей
def serialize_genus(genus, fields):
    diagnosis = genus.diagnosis
    row = {}

    # Основная информация
    if "Название рода" in fields:
        row["Название рода"] = genus.name or ""

    if "Полное название" in fields:
        row["Полное название"] = genus.full_name or ""

    if "Синонимы" in fields:
        row["Синонимы"] = ", ".join(s.name for s in genus.synonyms) if genus.synonyms else ""

    if "Типовой вид" in fields:
        row["Типовой вид"] = genus.type_species or ""

    if "Естественная принадлежность" in fields:
        row["Естественная принадлежность"] = genus.natural_affiliation or ""

    if "Сравнение" in fields:
        row["Сравнение"] = genus.comparison or ""

    if "Длина (мин)" in fields:
        row["Длина (мин)"] = genus.length_min or ""
    if "Длина (макс)" in fields:
        row["Длина (макс)"] = genus.length_max or ""
    if "Ширина (мин)" in fields:
        row["Ширина (мин)"] = genus.width_min or ""
    if "Ширина (макс)" in fields:
        row["Ширина (макс)"] = genus.width_max or ""

    # Диагноз
    if diagnosis:
        # Основные поля диагноза
        if "Инфратурма" in fields:
            row["Инфратурма"] = diagnosis.infraturma.name if diagnosis.infraturma else ""

        if "Характер щели разверзания" in fields:
            row["Характер щели разверзания"] = getattr(diagnosis, 'laesurae_character', "") or ""

        if "Наличие оторочки" in fields:
            row["Наличие оторочки"] = getattr(diagnosis, 'presence_of_rim', "") or ""

        if "Строение экзины" in fields:
            row["Строение экзины"] = ", ".join(
                structure.exine_structure for structure in getattr(diagnosis, 'exine_structure', [])
            ) if getattr(diagnosis, 'exine_structure', []) else ""

        if "Форма споры" in fields:
            row["Форма споры"] = diagnosis.form.name if diagnosis.form else ""

        if "Очертание" in fields:
            if hasattr(diagnosis, 'amb') and diagnosis.amb:
                amb_names = [a.amb for a in diagnosis.amb]
                row["Очертание"] = ", ".join(amb_names)
            else:
                row["Очертание"] = ""

        if "Форма сторон" in fields:
            if hasattr(diagnosis, 'sides_shape') and diagnosis.sides_shape:
                sides_names = [s.side_shape for s in diagnosis.sides_shape]
                row["Форма сторон"] = ", ".join(sides_names)
            else:
                row["Форма сторон"] = ""

        if "Форма углов" in fields:
            row["Форма углов"] = diagnosis.angles_shape.name if diagnosis.angles_shape else ""

        if "Форма щели разверзания" in fields:
            if hasattr(diagnosis, 'laesurae') and diagnosis.laesurae:
                laesurae_names = [l.laesurae_shape for l in diagnosis.laesurae]
                row["Форма щели разверзания"] = ", ".join(laesurae_names)
            else:
                row["Форма щели разверзания"] = ""

        if "Форма лучей щели" in fields:
            if hasattr(diagnosis, 'laesurae_rays') and diagnosis.laesurae_rays:
                rays_names = [r.rays_shape for r in diagnosis.laesurae_rays]
                row["Форма лучей щели"] = ", ".join(rays_names)
            else:
                row["Форма лучей щели"] = ""

        if "Длина лучей щели (мин)" in fields:
            row["Длина лучей щели (мин)"] = getattr(diagnosis, 'laesurae_rays_length_min', "")
        if "Длина лучей щели (макс)" in fields:
            row["Длина лучей щели (макс)"] = getattr(diagnosis, 'laesurae_rays_length_max', "")

        if "Выраженность ареа" in fields:
            row["Выраженность ареа"] = diagnosis.area_presence.name if diagnosis.area_presence else ""

        if "Форма контура споры" in fields:
            row["Форма контура споры"] = diagnosis.outline.name if diagnosis.outline else ""

        if "Причина неровности контура споры" in fields:
            row["Причина неровности контура споры"] = getattr(diagnosis, 'outline_irregularity_reason', "") or ""

        # Экзина
        if "Толщина экзины" in fields:
            if hasattr(diagnosis, 'exine_thickness') and diagnosis.exine_thickness:
                thickness_values = [
                    et.thickness.value for et in diagnosis.exine_thickness
                    if hasattr(et, 'thickness') and et.thickness
                ]
                row["Толщина экзины"] = ", ".join(thickness_values) if thickness_values else ""
            else:
                row["Толщина экзины"] = ""

        if "Структура экзины" in fields:
            if hasattr(diagnosis, 'exine_structure') and diagnosis.exine_structure:
                structure_names = [e.exine_structure for e in diagnosis.exine_structure]
                row["Структура экзины"] = ", ".join(structure_names)
            else:
                row["Структура экзины"] = ""

        # Форма разрастания экзины
        if hasattr(diagnosis, 'exine_growth_form') and diagnosis.exine_growth_form:
            exine_growth = diagnosis.exine_growth_form[0] if diagnosis.exine_growth_form else None
            if exine_growth:
                if "Тип" in fields:
                    row["Тип"] = exine_growth.exine_growth_type.name if exine_growth.exine_growth_type else ""

                if "Толщина" in fields:
                    row["Толщина"] = exine_growth.thickness.value if exine_growth.thickness else ""

                if "Ширина" in fields:
                    row["Ширина"] = exine_growth.width.value if exine_growth.width else ""

                if "Строение" in fields:
                    row["Строение"] = exine_growth.structure or ""
            else:
                if "Тип" in fields:
                    row["Тип"] = ""
                if "Толщина" in fields:
                    row["Толщина"] = ""
                if "Ширина" in fields:
                    row["Ширина"] = ""
                if "Строение" in fields:
                    row["Строение"] = ""

        # Экзоэкзина
        if hasattr(diagnosis, 'exoexine') and diagnosis.exoexine:
            exoexine = diagnosis.exoexine[0] if diagnosis.exoexine else None
            if exoexine:
                if "Экзоэкзина (толщина)" in fields:
                    row["Экзоэкзина (толщина)"] = exoexine.thickness.value if exoexine.thickness else ""

                if "Экзоэкзина (описание)" in fields:
                    row["Экзоэкзина (описание)"] = exoexine.description or ""
            else:
                if "Экзоэкзина (толщина)" in fields:
                    row["Экзоэкзина (толщина)"] = ""
                if "Экзоэкзина (описание)" in fields:
                    row["Экзоэкзина (описание)"] = ""

        # Интэкзина
        if hasattr(diagnosis, 'intexine') and diagnosis.intexine:
            intexine = diagnosis.intexine[0] if diagnosis.intexine else None
            if intexine:
                if "Интэкзина (толщина)" in fields:
                    row["Интэкзина (толщина)"] = intexine.thickness.value if intexine.thickness else ""

                if "Интэкзина (описание)" in fields:
                    row["Интэкзина (описание)"] = intexine.description or ""
            else:
                if "Интэкзина (толщина)" in fields:
                    row["Интэкзина (толщина)"] = ""
                if "Интэкзина (описание)" in fields:
                    row["Интэкзина (описание)"] = ""

        # Скульптура и орнаментация
        if "Скульптура" in fields:
            if hasattr(diagnosis, 'sculpture') and diagnosis.sculpture:
                sculptures = []
                for s in diagnosis.sculpture:
                    side = s.side.name if s.side else ""
                    sculpture = s.sculpture.sculpture if s.sculpture else ""
                    sculptures.append(f"{side}: {sculpture}")
                row["Скульптура"] = "; ".join(sculptures)
            else:
                row["Скульптура"] = ""

        if "Орнаментация" in fields:
            if hasattr(diagnosis, 'ornamentation') and diagnosis.ornamentation:
                ornaments = []
                for o in diagnosis.ornamentation:
                    side = o.side.name if o.side else ""
                    ornament = o.ornamentation.ornamentation if o.ornamentation else ""
                    ornaments.append(f"{side}: {ornament}")
                row["Орнаментация"] = "; ".join(ornaments)
            else:
                row["Орнаментация"] = ""

    stratigraphy_data = process_stratigraphy(genus, fields)
    row.update(stratigraphy_data)

    row.update(process_geography(genus, fields))
    # Виды
    if "Виды" in fields:
        row["Виды"] = ", ".join(s.name for s in genus.species) if genus.species else ""

    # return row
    return {k: v for k, v in row.items() if k in fields and v}


def serialize_species(species, fields):
    row = {
        "Название вида": species.name,
        "Старое название": species.old_name or "",
        "Источник": species.source or "",
        "Длина (мин)": species.length_min or "",
        "Длина (макс)": species.length_max or "",
        "Ширина (мин)": species.width_min or "",
        "Ширина (макс)": species.width_max or "",
    }

    stratigraphy_data = process_stratigraphy(species, fields)
    row.update(stratigraphy_data)

    row.update(process_geography(species, fields))

    return {k: v for k, v in row.items() if k in fields}


def export_data(data, fields, export_format, is_species=False):
    serializer = serialize_species if is_species else serialize_genus
    rows = [serializer(item, fields) for item in data]

    all_fields = set()
    for row in rows:
        all_fields.update(row.keys())

    ordered_fields = [f for f in fields if f in all_fields]
    df = pd.DataFrame(rows, columns=ordered_fields)

    save_to_file(df, export_format)
